import logging
import sqlite3
import re
import csv
import io
from datetime import datetime, date, time, timedelta
import requests
import smtplib
from email.message import EmailMessage
import threading

import matplotlib.pyplot as plt

from flask import (
    Flask,
    render_template,  # <-- endi render_template ishlatamiz
    request,
    redirect,
    url_for,
    session,
    send_file,
)


from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

# ===================== CONFIG =====================

DB_NAME = "finance_bot_full.db"
BASE_CURRENCY = "UZS"
FX_API_URL = "https://api.exchangerate.host/latest"

# Admin Telegram ID
ADMIN_TELEGRAM_ID = 1377933746

# Telegram bot token
TELEGRAM_BOT_TOKEN = "8216273158:AAGSATTvwAP_nio_3k-neVcX-plC3UmQN2k"

# Gmail SMTP sazlamalari
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = "qalmuratovazamat14@gmail.com"
SMTP_PASSWORD = "ptkv arcx hjws usvd"  # App parol
FROM_EMAIL = SMTP_USER

# Web admin panel login
ADMIN_WEB_USERNAME = "admin"
ADMIN_WEB_PASSWORD = "admin123"  # Paroli

# Web admin panel uchun Flask
admin_app = Flask(
    __name__,
    template_folder="templates",  # HTML fayllar papkasi
    static_folder="static",       # CSS, rasmlar, JS papkasi
)
admin_app.secret_key = "Proldi sir saqla"  # Ozgerttiriw kerek

# ===================== LOGGING =====================

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


# ===================== DATABASE =====================

def get_conn():
    return sqlite3.connect(DB_NAME)


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    # Paydalaniwshilar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            created_at TEXT,
            daily_reminder_enabled INTEGER DEFAULT 1,
            weekly_report_enabled INTEGER DEFAULT 1,
            email TEXT,
            daily_time TEXT,
            weekly_time TEXT,
            language TEXT DEFAULT 'UZB'
        )
        """
    )
    # Eski bazalar uchin til tirgegin qosiw
    try:
        cur.execute("ALTER TABLE users ADD COLUMN language TEXT DEFAULT 'UZB'")
    except sqlite3.OperationalError:
        pass

    # Hamyonlar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS wallets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            name TEXT,
            currency TEXT DEFAULT 'UZS',
            is_default INTEGER DEFAULT 0
        )
        """
    )

    # Kategoriyalar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            name TEXT,
            type TEXT,
            active INTEGER DEFAULT 1
        )
        """
    )

    # Tranzaksiyalar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            amount INTEGER,
            type TEXT,
            category_id INTEGER,
            note TEXT,
            created_at TEXT
        )
        """
    )

    # Maqsetler
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS goals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            name TEXT,
            target_amount INTEGER,
            saved_amount INTEGER DEFAULT 0,
            deadline TEXT
        )
        """
    )

    # Qarzlar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS debts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            direction TEXT,
            counterparty_name TEXT,
            total_amount INTEGER,
            remaining_amount INTEGER,
            created_at TEXT,
            due_date TEXT
        )
        """
    )

    # Byudjetler
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            category_id INTEGER,
            month INTEGER,
            year INTEGER,
            limit_amount INTEGER
        )
        """
    )

    # Shablonlar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            name TEXT,
            amount INTEGER,
            type TEXT,
            category TEXT,
            note TEXT
        )
        """
    )

    # Qaytalanatin tolewler
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS recurring_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            wallet_id INTEGER,
            amount INTEGER,
            type TEXT,
            category TEXT,
            note TEXT,
            day_of_month INTEGER,
            active INTEGER DEFAULT 1
        )
        """
    )

    # Quick kategoriya jagdayi
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_state (
            user_id INTEGER PRIMARY KEY,
            pending_category TEXT
        )
        """
    )

    # Hamyan tanlaw ushin kutilipatirgan tranzaksyalar
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pending_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            amount INTEGER,
            type TEXT,
            category TEXT,
            note TEXT
        )
        """
    )

    conn.commit()
    conn.close()


def ensure_user(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT user_id FROM users WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    if row is None:
        cur.execute(
            """
            INSERT INTO users (user_id, created_at, daily_time, weekly_time, language)
            VALUES (?, ?, ?, ?, ?)
            """,
            (user_id, datetime.utcnow().isoformat(), "22:00", "20:00", "UZB"),
        )
        conn.commit()
    conn.close()
    ensure_default_wallet(user_id)


def ensure_default_wallet(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM wallets WHERE user_id = ? AND is_default = 1",
        (user_id,),
    )
    row = cur.fetchone()
    if row is None:
        cur.execute(
            """
            INSERT INTO wallets (user_id, name, currency, is_default)
            VALUES (?, ?, ?, 1)
            """,
            (user_id, "Asosiy hamyon", BASE_CURRENCY),
        )
        conn.commit()
    conn.close()


def list_wallets_db(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, currency, is_default
        FROM wallets
        WHERE user_id = ?
        ORDER BY id
        """,
        (user_id,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def add_wallet_db(user_id: int, name: str, currency: str = BASE_CURRENCY):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO wallets (user_id, name, currency, is_default)
        VALUES (?, ?, ?, 0)
        """,
        (user_id, name, currency),
    )
    conn.commit()
    conn.close()


def set_default_wallet_db(user_id: int, wallet_id: int) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM wallets WHERE id = ? AND user_id = ?",
        (wallet_id, user_id),
    )
    row = cur.fetchone()
    if row is None:
        conn.close()
        return False
    cur.execute("UPDATE wallets SET is_default = 0 WHERE user_id = ?", (user_id,))
    cur.execute(
        "UPDATE wallets SET is_default = 1 WHERE id = ?",
        (wallet_id,),
    )
    conn.commit()
    conn.close()
    return True


def get_default_wallet_id(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM wallets WHERE user_id = ? AND is_default = 1",
        (user_id,),
    )
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None


def get_or_create_category(user_id: int, wallet_id: int, name: str, ttype: str):
    name = name.strip().lower()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id FROM categories
        WHERE user_id = ? AND wallet_id = ? AND name = ? AND type = ? AND active = 1
        """,
        (user_id, wallet_id, name, ttype),
    )
    row = cur.fetchone()
    if row:
        cid = row[0]
        conn.close()
        return cid
    cur.execute(
        """
        INSERT INTO categories (user_id, wallet_id, name, type, active)
        VALUES (?, ?, ?, ?, 1)
        """,
        (user_id, wallet_id, name, ttype),
    )
    conn.commit()
    cid = cur.lastrowid
    conn.close()
    return cid


def list_categories_db(user_id: int, wallet_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, type
        FROM categories
        WHERE user_id = ? AND wallet_id = ? AND active = 1
        ORDER BY name
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def deactivate_category_db(user_id: int, wallet_id: int, name: str) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id FROM categories
        WHERE user_id = ? AND wallet_id = ? AND name = ? AND active = 1
        """,
        (user_id, wallet_id, name.lower()),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    cur.execute(
        "UPDATE categories SET active = 0 WHERE id = ?",
        (row[0],),
    )
    conn.commit()
    conn.close()
    return True


def rename_category_db(user_id: int, wallet_id: int, old: str, new: str) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id FROM categories
        WHERE user_id = ? AND wallet_id = ? AND name = ? AND active = 1
        """,
        (user_id, wallet_id, old.lower()),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    cur.execute(
        "UPDATE categories SET name = ? WHERE id = ?",
        (new.lower(), row[0]),
    )
    conn.commit()
    conn.close()
    return True


def add_transaction_db(
    user_id: int,
    wallet_id: int,
    amount: int,
    ttype: str,
    category_name: str,
    note: str = "",
    ):
    category_id = get_or_create_category(user_id, wallet_id, category_name, ttype)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO transactions (user_id, wallet_id, amount, type,
                                  category_id, note, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            wallet_id,
            amount,
            ttype,
            category_id,
            note,
            datetime.utcnow().isoformat(),
        ),
    )
    conn.commit()
    conn.close()
    return category_id


def get_period_stats(user_id: int, wallet_id: int, start_dt: datetime, end_dt: datetime):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT type, SUM(amount)
        FROM transactions
        WHERE user_id = ?
          AND wallet_id = ?
          AND datetime(created_at) >= ?
          AND datetime(created_at) <= ?
        GROUP BY type
        """,
        (user_id, wallet_id, start_dt.isoformat(), end_dt.isoformat()),
    )
    income = 0
    expense = 0
    for ttype, total in cur.fetchall():
        if ttype == "income":
            income = total or 0
        elif ttype == "expense":
            expense = total or 0

    cur.execute(
        """
        SELECT c.name, SUM(t.amount) as total
        FROM transactions t
        JOIN categories c ON t.category_id = c.id
        WHERE t.user_id = ?
          AND t.wallet_id = ?
          AND t.type = 'expense'
          AND datetime(t.created_at) >= ?
          AND datetime(t.created_at) <= ?
        GROUP BY c.name
        ORDER BY total DESC
        """,
        (user_id, wallet_id, start_dt.isoformat(), end_dt.isoformat()),
    )
    cat_rows = cur.fetchall()
    conn.close()
    return income, expense, cat_rows


def get_wallet_balance_db(user_id: int, wallet_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT type, SUM(amount)
        FROM transactions
        WHERE user_id = ? AND wallet_id = ?
        GROUP BY type
        """,
        (user_id, wallet_id),
    )
    income = 0
    expense = 0
    for ttype, total in cur.fetchall():
        if ttype == "income":
            income = total or 0
        elif ttype == "expense":
            expense = total or 0
    conn.close()
    return (income or 0) - (expense or 0)


def set_budget_db(
    user_id: int,
    wallet_id: int,
    category_name: str,
    limit_amount: int,
    year: int,
    month: int,
):
    cat_id = get_or_create_category(user_id, wallet_id, category_name, "expense")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id FROM budgets
        WHERE user_id = ? AND wallet_id = ? AND category_id = ?
          AND year = ? AND month = ?
        """,
        (user_id, wallet_id, cat_id, year, month),
    )
    row = cur.fetchone()
    if row:
        cur.execute(
            "UPDATE budgets SET limit_amount = ? WHERE id = ?",
            (limit_amount, row[0]),
        )
    else:
        cur.execute(
            """
            INSERT INTO budgets (user_id, wallet_id, category_id, month, year, limit_amount)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (user_id, wallet_id, cat_id, month, year, limit_amount),
        )
    conn.commit()
    conn.close()


def get_budget_status_for_category(
    user_id: int, wallet_id: int, category_id: int, year: int, month: int
):
    start_dt = datetime(year, month, 1)
    if month == 12:
        end_dt = datetime(year + 1, 1, 1) - timedelta(seconds=1)
    else:
        end_dt = datetime(year, month + 1, 1) - timedelta(seconds=1)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT limit_amount FROM budgets
        WHERE user_id = ? AND wallet_id = ? AND category_id = ?
          AND year = ? AND month = ?
        """,
        (user_id, wallet_id, category_id, year, month),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return None
    limit_amount = row[0]

    cur.execute(
        """
        SELECT SUM(amount)
        FROM transactions
        WHERE user_id = ?
          AND wallet_id = ?
          AND category_id = ?
          AND type = 'expense'
          AND datetime(created_at) >= ?
          AND datetime(created_at) <= ?
        """,
        (
            user_id,
            wallet_id,
            category_id,
            start_dt.isoformat(),
            end_dt.isoformat(),
        ),
    )
    spent_row = cur.fetchone()
    spent = spent_row[0] or 0
    conn.close()
    return limit_amount, spent


def get_all_budgets_status(user_id: int, wallet_id: int, year: int, month: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT b.id, b.category_id, b.limit_amount, c.name
        FROM budgets b
        JOIN categories c ON b.category_id = c.id
        WHERE b.user_id = ? AND b.wallet_id = ? AND b.year = ? AND b.month = ?
        """,
        (user_id, wallet_id, year, month),
    )
    rows = cur.fetchall()
    conn.close()
    results = []
    for _, cat_id, limit_amount, cat_name in rows:
        status = get_budget_status_for_category(user_id, wallet_id, cat_id, year, month)
        if status:
            limit_amount, spent = status
            results.append((cat_name, limit_amount, spent))
    return results


def add_goal_db(user_id: int, wallet_id: int, name: str, target_amount: int, deadline: str | None):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO goals (user_id, wallet_id, name, target_amount, deadline)
        VALUES (?, ?, ?, ?, ?)
        """,
        (user_id, wallet_id, name, target_amount, deadline),
    )
    conn.commit()
    conn.close()


def list_goals_db(user_id: int, wallet_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, target_amount, saved_amount, deadline
        FROM goals
        WHERE user_id = ? AND wallet_id = ?
        ORDER BY id
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def add_to_goal_db(user_id: int, wallet_id: int, goal_id: int, amount: int) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT saved_amount FROM goals
        WHERE id = ? AND user_id = ? AND wallet_id = ?
        """,
        (goal_id, user_id, wallet_id),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    saved = row[0] or 0
    new_saved = saved + amount
    cur.execute(
        "UPDATE goals SET saved_amount = ? WHERE id = ?",
        (new_saved, goal_id),
    )
    conn.commit()
    conn.close()
    return True


def find_goal_by_name(user_id: int, wallet_id: int, name_part: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, target_amount, saved_amount
        FROM goals
        WHERE user_id = ?
          AND wallet_id = ?
          AND LOWER(name) LIKE ?
        ORDER BY id
        """,
        (user_id, wallet_id, f"%{name_part.lower()}%",),
    )
    row = cur.fetchone()
    conn.close()
    return row


def add_debt_db(
    user_id: int,
    wallet_id: int,
    direction: str,
    name: str,
    amount: int,
    due_date: str | None,
):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO debts (user_id, wallet_id, direction, counterparty_name,
                           total_amount, remaining_amount, created_at, due_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_id,
            wallet_id,
            direction,
            name,
            amount,
            amount,
            datetime.utcnow().isoformat(),
            due_date,
        ),
    )
    conn.commit()
    conn.close()


def list_debts_db(user_id: int, wallet_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, direction, counterparty_name, total_amount, remaining_amount, due_date
        FROM debts
        WHERE user_id = ? AND wallet_id = ? AND remaining_amount > 0
        ORDER BY id
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def pay_debt_db(user_id: int, wallet_id: int, debt_id: int, amount: int):
    """
    True/False ham qalgan summani qaytaradi (yaki None).
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT remaining_amount FROM debts
        WHERE id = ? AND user_id = ? AND wallet_id = ?
        """,
        (debt_id, user_id, wallet_id),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return False, None
    remaining = row[0]
    new_remaining = max(0, remaining - amount)
    cur.execute(
        "UPDATE debts SET remaining_amount = ? WHERE id = ?",
        (new_remaining, debt_id),
    )
    conn.commit()
    conn.close()
    return True, new_remaining


def export_transactions_csv(user_id: int, wallet_id: int) -> io.BytesIO:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT t.created_at, t.type, t.amount, c.name, t.note, w.name, w.currency
        FROM transactions t
        JOIN categories c ON t.category_id = c.id
        JOIN wallets w ON t.wallet_id = w.id
        WHERE t.user_id = ? AND t.wallet_id = ?
        ORDER BY datetime(t.created_at)
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["date", "type", "amount", "category", "note", "wallet", "currency"])
    for created_at, ttype, amount, cat_name, note, wallet_name, currency in rows:
        writer.writerow([created_at, ttype, amount, cat_name, note or "", wallet_name, currency])

    mem = io.BytesIO()
    mem.write(output.getvalue().encode("utf-8"))
    mem.seek(0)
    return mem


def export_transactions_excel(user_id: int, wallet_id: int) -> io.BytesIO:
    import openpyxl
    from openpyxl.utils import get_column_letter

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT t.created_at, t.type, t.amount, c.name, t.note, w.name, w.currency
        FROM transactions t
        JOIN categories c ON t.category_id = c.id
        JOIN wallets w ON t.wallet_id = w.id
        WHERE t.user_id = ? AND t.wallet_id = ?
        ORDER BY datetime(t.created_at)
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["date", "type", " amount", "category", "note", "wallet", "currency"]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return mem


def add_template_db(user_id: int, wallet_id: int, name: str, amount: int, ttype: str, category: str, note: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO templates (user_id, wallet_id, name, amount, type, category, note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (user_id, wallet_id, name, amount, ttype, category, note),
    )
    conn.commit()
    conn.close()


def list_templates_db(user_id: int, wallet_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, amount, type, category, note
        FROM templates
        WHERE user_id = ? AND wallet_id = ?
        ORDER BY id
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_template_db(user_id: int, wallet_id: int, template_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, amount, type, category, note
        FROM templates
        WHERE user_id = ? AND wallet_id = ? AND id = ?
        """,
        (user_id, wallet_id, template_id),
    )
    row = cur.fetchone()
    conn.close()
    return row


def add_recurring_db(
    user_id: int,
    wallet_id: int,
    amount: int,
    ttype: str,
    category: str,
    note: str,
    day_of_month: int,
):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO recurring_payments
          (user_id, wallet_id, amount, type, category, note, day_of_month, active)
        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (user_id, wallet_id, amount, ttype, category, note, day_of_month),
    )
    conn.commit()
    conn.close()


def list_recurring_db(user_id: int, wallet_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, amount, type, category, note, day_of_month, active
        FROM recurring_payments
        WHERE user_id = ? AND wallet_id = ?
        ORDER BY day_of_month
        """,
        (user_id, wallet_id),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def deactivate_recurring_db(user_id: int, wallet_id: int, rid: int) -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id FROM recurring_payments
        WHERE id = ? AND user_id = ? AND wallet_id = ?
        """,
        (rid, user_id, wallet_id),
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    cur.execute(
        "UPDATE recurring_payments SET active = 0 WHERE id = ?",
        (rid,),
    )
    conn.commit()
    conn.close()
    return True


def set_email_db(user_id: int, email: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET email = ? WHERE user_id = ?",
        (email, user_id),
    )
    conn.commit()
    conn.close()


def get_email_db(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT email FROM users WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None


def get_user_notification_times(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT daily_time, weekly_time FROM users WHERE user_id = ?",
        (user_id,),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return "22:00", "20:00"
    return row[0] or "22:00", row[1] or "20:00"


def set_daily_time_db(user_id: int, tstr: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET daily_time = ? WHERE user_id = ?",
        (tstr, user_id),
    )
    conn.commit()
    conn.close()


def set_weekly_time_db(user_id: int, tstr: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET weekly_time = ? WHERE user_id = ?",
        (tstr, user_id),
    )
    conn.commit()
    conn.close()


def get_pending_category_db(user_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT pending_category FROM user_state WHERE user_id = ?",
        (user_id,),
    )
    row = cur.fetchone()
    conn.close()
    return row[0] if row and row[0] else None


def set_pending_category_db(user_id: int, category: str | None):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT user_id FROM user_state WHERE user_id = ?",
        (user_id,),
    )
    row = cur.fetchone()
    if row:
        cur.execute(
            "UPDATE user_state SET pending_category = ? WHERE user_id = ?",
            (category, user_id),
        )
    else:
        cur.execute(
            "INSERT INTO user_state (user_id, pending_category) VALUES (?, ?)",
            (user_id, category),
        )
    conn.commit()
    conn.close()


def create_pending_transaction_db(user_id: int, amount: int, ttype: str, category: str, note: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO pending_transactions (user_id, amount, type, category, note)
        VALUES (?, ?, ?, ?, ?)
        """,
        (user_id, amount, ttype, category, note),
    )
    conn.commit()
    pid = cur.lastrowid
    conn.close()
    return pid


def get_pending_transaction_db(user_id: int, pending_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, user_id, amount, type, category, note
        FROM pending_transactions
        WHERE id = ? AND user_id = ?
        """,
        (pending_id, user_id),
    )
    row = cur.fetchone()
    conn.close()
    return row


def delete_pending_transaction_db(pending_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM pending_transactions WHERE id = ?", (pending_id,))
    conn.commit()
    conn.close()


# ---- Til funksiyalari ----

def get_user_language(user_id: int) -> str:
    """
    Paydalanuwshi tili (UZB / QQR).
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT language FROM users WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    if not row or not row[0]:
        return "UZB"
    return row[0]


def set_user_language_db(user_id: int, lang: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET language = ? WHERE user_id = ?",
        (lang, user_id),
    )
    conn.commit()
    conn.close()


# ===================== UTIL FUNKSIYALAR =====================

def main_menu_keyboard(lang: str = "UZB"):

    # Agar QQR varianti yasamoqchi bo'lsang, shu yerda shart qo'yib yozasan:
    # if lang == "QQR":
    #     btn_add = "➕ Kiris/Shigys"  # misol, o'zing yozib olasan
    # else:
    #     btn_add = "➕ Daromad/Xarajat"
    btn_add = "➕ Daramat/Shigim"
    btn_today = "📊 Bugun"
    btn_month = "📆 Ay"
    btn_goals = "🎯 Maqsetler"
    btn_settings = "⚙️ Sazlamalar"

    keyboard = [
        [KeyboardButton(btn_add)],
        [KeyboardButton(btn_today), KeyboardButton(btn_month)],
        [KeyboardButton(btn_goals), KeyboardButton(btn_settings)],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def quick_categories_keyboard():
    buttons = [
        [
            InlineKeyboardButton("🍽 Awqat", callback_data="cat:ovqat"),
            InlineKeyboardButton("🚍 Transport", callback_data="cat:transport"),
        ],
        [
            InlineKeyboardButton("🏠 Ijara", callback_data="cat:ijara"),
            InlineKeyboardButton("☕️ Kafe", callback_data="cat:kafe"),
        ],
        [
            InlineKeyboardButton("🛒 Bazar", callback_data="cat:bozor"),
            InlineKeyboardButton("🛍 Kiyim", callback_data="cat:kiyim"),
        ],
    ]
    return InlineKeyboardMarkup(buttons)


def get_week_bounds(today: date):
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    start_dt = datetime(start.year, start.month, start.day)
    end_dt = datetime(end.year, end.month, end.day, 23, 59, 59)
    return start_dt, end_dt


def parse_transaction_text(text: str):
    match = re.search(r"[+-]?\d+", text)
    if not match:
        return None

    amount_str = match.group()
    try:
        amount = int(amount_str)
    except ValueError:
        return None

    if amount_str.startswith("+"):
        ttype = "income"
    elif amount_str.startswith("-"):
        ttype = "expense"
        amount = abs(amount)
    else:
        ttype = "expense"

    after = text[match.end():].strip()
    if not after:
        category = "basqa"
        note = ""
    else:
        parts = after.split(maxsplit=1)
        category = parts[0].strip().lower()
        note = parts[1].strip() if len(parts) == 2 else ""

    return amount, ttype, category, note


def parse_goal_contribution_text(text: str):
    lower = text.lower()
    if "goal" not in lower:
        return None
    amount_match = re.search(r"\d+", text)
    if not amount_match:
        return None
    amount = int(amount_match.group())
    idx = lower.find("goal")
    goal_part = text[idx + len("goal"):].strip()
    if not goal_part:
        return None
    return amount, goal_part


def convert_to_base_currency(amount: int, currency: str):
    if currency == BASE_CURRENCY:
        return amount
    try:
        resp = requests.get(
            FX_API_URL,
            params={"base": currency, "symbols": BASE_CURRENCY},
            timeout=5,
        )
        if resp.status_code != 200:
            return None
        data = resp.json()
        rate = data.get("rates", {}).get(BASE_CURRENCY)
        if not rate:
            return None
        uzs_amount = int(amount * rate)
        return uzs_amount
    except Exception as e:
        logger.warning(f"FX API error for {currency}: {e}")
        return None


def parse_time_str(tstr: str):
    try:
        hh, mm = tstr.split(":")
        hh = int(hh)
        mm = int(mm)
        if not (0 <= hh <= 23 and 0 <= mm <= 59):
            return None
        return time(hour=hh, minute=mm)
    except Exception:
        return None


def build_chart_image(user_id: int, wallet_id: int, start_dt: datetime, end_dt: datetime):
    income, expense, cat_rows = get_period_stats(user_id, wallet_id, start_dt, end_dt)
    if not cat_rows or expense == 0:
        return None
    labels = [name for name, total in cat_rows]
    sizes = [total for name, total in cat_rows]

    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels, autopct="%1.1f%%")
    ax.set_title("Kategoriyalar boyinsha shigimlar")

    buf = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf


def quick_category_title(key: str):
    mapping = {
        "awqat": "Awqat",
        "transport": "Transport",
        "ijara": "Ijara",
        "kafe": "Kafe",
        "bazar": "Bazar",
        "kiyim": "Kiyim",
    }
    return mapping.get(key, key)


# ===================== JOB SCHEDULING HELPERS =====================

def schedule_daily_job(job_queue, user_id: int, tstr: str):
    t = parse_time_str(tstr) or time(hour=22, minute=0)
    name = f"daily_{user_id}"
    for job in job_queue.get_jobs_by_name(name):
        job.schedule_removal()
    job_queue.run_daily(
        daily_reminder_job,
        time=t,
        days=(0, 1, 2, 3, 4, 5, 6),
        data={"user_id": user_id},
        name=name,
    )


def schedule_weekly_job(job_queue, user_id: int, tstr: str):
    t = parse_time_str(tstr) or time(hour=20, minute=0)
    name = f"weekly_{user_id}"
    for job in job_queue.get_jobs_by_name(name):
        job.schedule_removal()
    job_queue.run_daily(
        weekly_report_job,
        time=t,
        days=(6,),
        data={"user_id": user_id},
        name=name,
    )


def schedule_user_jobs(job_queue, user_id: int):
    d, w = get_user_notification_times(user_id)
    schedule_daily_job(job_queue, user_id, d)
    schedule_weekly_job(job_queue, user_id, w)


def schedule_all_users_jobs(job_queue):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT user_id, daily_time, weekly_time FROM users")
    rows = cur.fetchall()
    conn.close()
    for user_id, d, w in rows:
        schedule_daily_job(job_queue, user_id, d or "22:00")
        schedule_weekly_job(job_queue, user_id, w or "20:00")


# ===================== HANDLERLAR =====================

async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    schedule_user_jobs(context.application.job_queue, user.id)
    lang = get_user_language(user.id)

    text = (
        "Assalomu alaykum! 👋\n\n"
        "Bul bot jardeminde shaxsiy mbudjetinizdi basqariwiniz mumkin.\n\n"
        "Shigim/daramad qosiw:\n"
        "  - `-50000 awqat`\n"
        "  - `+200000 ayliq`\n"
        "  - `bugun 120000 bazar`\n\n"
        "Tiykargi komandalar:\n"
        "  /today - bugun\n"
        "  /week - hapte\n"
        "  /month - oy\n"
        "  /year - jil\n"
        "  /report - dawir tanlaw\n\n"
        "Hamyanlar:\n"
        "  /wallets, /addwallet, /setwallet, /totalwealth\n\n"
        "Kategoriyalar ham byudjet:\n"
        "  /categories, /setbudget, /budget\n\n"
        "Maqsetler:\n"
        "  /goals, /addgoal, /addgoalsum\n"
        "  yaki: `+200000 goal telefon ushin`\n\n"   # ✅ backtick yopildi
        "Qarzlar:\n"
        "  /debts, /lend, /borrow, /paydebt\n"
        "  yaki: \"Men bankka bugun 500000 toledim kredit uchun\" (auto qarz tolewi)\n\n"
        "Shablon ham takrarlaniwshi tolewler ushun:\n"
        "  /templates, /addtemplate, /usetemplate, /addrecurring, /recurring, /delrecurring\n\n"
        "Eksport:\n"
        "  /exportcsv, /exportxlsx, /setemail, /emailreport\n\n"
        "Sazlamalar:\n"
        "  /settings, /toggledaily, /toggleweekly, /setdailytime, /setweeklytime\n\n"
        "Tildi ozgerttiriw: /lang (UZB / QQR)\n"
    )

    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=main_menu_keyboard(lang),
    )


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start_cmd(update, context)


# --- Til tanlash komandasi ---

async def lang_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    keyboard = [
        [InlineKeyboardButton("🇺🇿 O'zbek tili (UZB)", callback_data="lang:UZB")],
        [InlineKeyboardButton("qqr Qaraqalpaq tili (QQR)", callback_data="lang:QQR")],
    ]
    await update.message.reply_text(
        "Tilni tanlang:",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def lang_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data  # lang:UZB / lang:QQR
    _, lang = data.split(":", 1)
    user_id = query.from_user.id
    ensure_user(user_id)
    if lang not in ("UZB", "QQR"):
        await query.message.reply_text("Qate til kodi.")
        return
    set_user_language_db(user_id, lang)
    if lang == "UZB":
        msg = "Til O'zbek tiliga o'zgartirildi ✅"
    else:
        msg = "Til Qaraqalpaq tiline ozgerttirildi ✅"
    await query.message.reply_text(
        msg,
        reply_markup=main_menu_keyboard(lang),
    )


# --- Hamyan komandalar ---

async def wallets_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    rows = list_wallets_db(user.id)
    if not rows:
        await update.message.reply_text("Sizde ele hamyan joq. /addwallet arqali qosiwiniz mumkin.")
        return
    lines = ["Sizdin hamyanlariniz:"]
    for wid, name, currency, is_default in rows:
        mark = " (joriy)" if is_default else ""
        lines.append(f"ID {wid}: {name} [{currency}]{mark}")
    lines.append("\nHazirgi hamyandi almastiriw: /setwallet ID")
    await update.message.reply_text("\n".join(lines))


async def addwallet_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    args = context.args
    if not args:
        await update.message.reply_text(
            "Paydalaniw: /addwallet Ati [Valyuta]\nMisalga: /addwallet Naq UZS"
        )
        return
    if len(args) == 1:
        name = args[0]
        currency = BASE_CURRENCY
    else:
        name = " ".join(args[:-1])
        currency = args[-1].upper()
    add_wallet_db(user.id, name, currency)
    await update.message.reply_text(f"Jana hamyan qosildi ✅\nAti: {name}, Valyuta: {currency}")


async def setwallet_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    if not context.args:
        await update.message.reply_text("Paydalaniw: /setwallet ID")
        return
    try:
        wid = int(context.args[0])
    except ValueError:
        await update.message.reply_text("ID putun san boliwi kerek.")
        return
    ok = set_default_wallet_db(user.id, wid)
    if not ok:
        await update.message.reply_text("Bunday ID li hamyan tabilmadi.")
    else:
        await update.message.reply_text(f"Hazirgi hamyan ID {wid} ga ozgerttirildi ✅")


async def totalwealth_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallets = list_wallets_db(user.id)
    if not wallets:
        await update.message.reply_text("Hamyanlar tawilmadi.")
        return

    lines = []
    total_base = 0
    trouble = []
    for wid, name, currency, is_default in wallets:
        balance = get_wallet_balance_db(user.id, wid)
        converted = convert_to_base_currency(balance, currency)
        if converted is None:
            trouble.append(currency)
            lines.append(f"{name} [{currency}]: {balance} (kurs alinbadi)")
        else:
            total_base += converted
            if currency == BASE_CURRENCY:
                lines.append(f"{name} [{currency}]: {balance}")
            else:
                lines.append(f"{name} [{currency}]: {balance} ≈ {converted} {BASE_CURRENCY}")

    lines.append(f"\nUliwmaliq bayliq: {total_base} {BASE_CURRENCY}")
    if trouble:
        lines.append(f"(Valyuta kursinda qatelik: {', '.join(set(trouble))})")
    await update.message.reply_text("\n".join(lines))


# --- Statistika komandalar ---

async def stats_period_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE, period: str, via_callback: bool = False):
    if via_callback:
        query = update.callback_query
        user_id = query.from_user.id
        reply_func = query.edit_message_text
        chat_id = query.message.chat.id
    else:
        user = update.effective_user
        user_id = user.id
        reply_func = update.message.reply_text
        chat_id = update.effective_chat.id

    ensure_user(user_id)
    wallet_id = get_default_wallet_id(user_id)
    if wallet_id is None:
        await reply_func("Hamyan tawilmadi. /start ti basip korin.")
        return

    today = date.today()
    if period == "today":
        start_dt = datetime(today.year, today.month, today.day)
        end_dt = datetime(today.year, today.month, today.day, 23, 59, 59)
        title = "Bugungi statistika"
    elif period == "week":
        start_dt, end_dt = get_week_bounds(today)
        title = "Usi hapte statistikasi"
    elif period == "month":
        start_dt = datetime(today.year, today.month, 1)
        if today.month == 12:
            end_dt = datetime(today.year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end_dt = datetime(today.year, today.month + 1, 1) - timedelta(seconds=1)
        title = "Usi ay statistikasi"
    elif period == "year":
        start_dt = datetime(today.year, 1, 1)
        end_dt = datetime(today.year, 12, 31, 23, 59, 59)
        title = "Usi jil statistikasi"
    else:
        return

    income, expense, cat_rows = get_period_stats(user_id, wallet_id, start_dt, end_dt)
    balance = (income or 0) - (expense or 0)

    text = f"📊 {title}:\n\nDaramad: {income or 0} som\nShigim: {expense or 0} som\nBalans: {balance} som\n"

    if expense > 0 and cat_rows:
        text += "\nKategoriyalar kesiminde shigimlar:\n"
        for name, total in cat_rows:
            percent = int((total * 100) / expense)
            text += f"  • {name}: {total} som ({percent}%)\n"

    if period == "month":
        if today.month == 1:
            prev_year = today.year - 1
            prev_month = 12
        else:
            prev_year = today.year
            prev_month = today.month - 1
        prev_start = datetime(prev_year, prev_month, 1)
        if prev_month == 12:
            prev_end = datetime(prev_year + 1, 1, 1) - timedelta(seconds=1)
        else:
            prev_end = datetime(prev_year, prev_month + 1, 1) - timedelta(seconds=1)

        p_inc, p_exp, _ = get_period_stats(user_id, wallet_id, prev_start, prev_end)
        if p_exp > 0:
            change = expense - p_exp
            perc = int((change * 100) / p_exp)
            if change > 0:
                text += f"\nOtken ayga salistirganda shigimlar {change} somga (+{perc}%) kobeygen."
            elif change < 0:
                text += f"\nOtken ayga salistirganda shigimlar {-change} somga ({perc}%) azaygan."
            else:
                text += "\nOtken ayga salistirganda shigimlar ozgermegen."
        else:
            text += "\nOtken ay uchun magliwmat az, salistiriw imkansiz."

    await reply_func(text)

    chart_buf = build_chart_image(user_id, wallet_id, start_dt, end_dt)
    if chart_buf is not None:
        try:
            await context.bot.send_photo(
                chat_id=chat_id,
                photo=chart_buf,
                caption="Kategoriyalar boyinsha shigimlar diagrammasi 🥧",
            )
        except Exception as e:
            logger.warning(f"Chart send error: {e}")


async def today_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await stats_period_cmd(update, context, "today")


async def week_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await stats_period_cmd(update, context, "week")


async def month_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await stats_period_cmd(update, context, "month")


async def year_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await stats_period_cmd(update, context, "year")


async def report_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [
            InlineKeyboardButton("Bugun", callback_data="report:today"),
            InlineKeyboardButton("Hapte", callback_data="report:week"),
        ],
        [
            InlineKeyboardButton("Ay", callback_data="report:month"),
            InlineKeyboardButton("Jil", callback_data="report:year"),
        ],
    ]
    await update.message.reply_text(
        "Qaysi dawir uchun esabat kerek?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def report_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("report:"):
        period = data.split(":", 1)[1]
        await stats_period_cmd(update, context, period, via_callback=True)


# --- Kategoriyalar / byudjet ---

async def categories_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return

    if not context.args:
        rows = list_categories_db(user.id, wallet_id)
        if not rows:
            await update.message.reply_text("Ele kategoriyalar joq. shigimlardi qossaniz, avtomatik jaratiladi.")
            return
        lines = ["Kategoriyalar:"]
        for cid, name, ttype in rows:
            icon = "➕" if ttype == "income" else "➖"
            lines.append(f"{icon} {name}")
        lines.append("\nQosiw: /categories add KATEGORIYA")
        lines.append("Oshiriw: /categories del KATEGORIYA")
        lines.append("Atin ozgertiriw: /categories rename ESKI JANA\nMisalga: /categories rename awqatlaniw ushin")
        await update.message.reply_text("\n".join(lines))
        return

    sub = context.args[0].lower()
    if sub == "add" and len(context.args) >= 2:
        name = context.args[1].lower()
        get_or_create_category(user.id, wallet_id, name, "expense")
        await update.message.reply_text(f"Kategoriya qosildi: {name} ✅")
    elif sub == "del" and len(context.args) >= 2:
        name = context.args[1].lower()
        ok = deactivate_category_db(user.id, wallet_id, name)
        if ok:
            await update.message.reply_text(f"Kategoriya ochirildi (aktiv emes): {name} ✅")
        else:
            await update.message.reply_text("Bunday kategoriya topilmadi.")
    elif sub == "rename" and len(context.args) >= 3:
        old = context.args[1].lower()
        new = context.args[2].lower()
        ok = rename_category_db(user.id, wallet_id, old, new)
        if ok:
            await update.message.reply_text(f"Kategoriya ati '{old}' dan '{new}' ga ozgerttirildi ✅")
        else:
            await update.message.reply_text("Bunday kategoriya tawilmadi.")
    else:
        await update.message.reply_text(
            "Paydalaniw:\n"
            "/categories - dizim\n"
            "/categories add KATEGORIYA\n"
            "/categories del KATEGORIYA\n"
            "/categories rename OLD NEW"
        )


async def setbudget_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Paydalaniw: /setbudget SUMMA KATEGORIYA\nMisalga: /setbudget 1200000 awqat"
        )
        return
    try:
        amount = int(context.args[0])
    except ValueError:
        await update.message.reply_text("SUMMA putun san boliwi kerek.")
        return
    category = " ".join(context.args[1:]).lower()
    today = date.today()
    set_budget_db(user.id, wallet_id, category, amount, today.year, today.month)
    await update.message.reply_text(
        f"Byudjet ornatildi ✅\nKategoriya: {category}\nLimit: {amount} som (ayina)"
    )


async def budget_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    today = date.today()
    rows = get_all_budgets_status(user.id, wallet_id, today.year, today.month)
    if not rows:
        await update.message.reply_text("Hazirgi ay uchun byudjetler ornatilmagan.")
        return
    lines = [f"📆 {today.month}-{today.year} byudjetleri:"]
    for cname, limit_amount, spent in rows:
        percent = int(spent * 100 / limit_amount) if limit_amount > 0 else 0
        lines.append(f"• {cname}: {spent}/{limit_amount} som ({percent}%)")
    await update.message.reply_text("\n".join(lines))


async def check_budget_alert_and_reply(
    reply_func,
    user_id: int,
    wallet_id: int,
    category_id: int,
    category_name: str,
):
    today = date.today()
    status = get_budget_status_for_category(user_id, wallet_id, category_id, today.year, today.month)
    if not status:
        return
    limit_amount, spent = status
    if limit_amount <= 0:
        return
    percent = int(spent * 100 / limit_amount)
    if percent >= 100:
        await reply_func(
            f"⚠️ Diqqat! '{category_name}' kategoriyasi uchun byudjet 100% ten asip ketti.\n"
            f"Sariplaniwlar: {spent}/{limit_amount} som ({percent}%)"
        )
    elif percent >= 80:
        await reply_func(
            f"ℹ️ Eskertiw: '{category_name}' kategoriyasi uchun byudjettin {percent}% sarpandi.\n"
            f"Sarplangan: {spent}/{limit_amount} som"
        )


# --- Goals ---

async def addgoal_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Paydalaniw: /addgoal SUMMA ATI [YYYY-MM-DD]\n"
            "Masalga: /addgoal 5000000 Telefon ushi 2025-12-01"
        )
        return
    try:
        amount = int(context.args[0])
    except ValueError:
        await update.message.reply_text("SUMMA putun san boliwi kerek.")
        return
    if len(context.args) >= 3:
        name = " ".join(context.args[1:-1])
        deadline = context.args[-1]
    else:
        name = context.args[1]
        deadline = None
    add_goal_db(user.id, wallet_id, name, amount, deadline)
    await update.message.reply_text(
        f"Maqset qosildi ✅\nAti: {name}\nSumma: {amount} som\nDeadline: {deadline or 'joq'}"
    )


async def goals_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    rows = list_goals_db(user.id, wallet_id)
    if not rows:
        await update.message.reply_text(
            "Sizda ele maqsetler joq.\nJaratiw: /addgoal SUMMA ATI [YYYY-MM-DD]"
        )
        return
    lines = ["Sizdin maqsetleriniz:"]
    today = date.today()
    for gid, name, target, saved, deadline in rows:
        saved = saved or 0
        percent = int(saved * 100 / target) if target > 0 else 0
        line = f"ID {gid}: {name} – {saved}/{target} som ({percent}%)"
        if deadline:
            try:
                d = datetime.fromisoformat(deadline).date()
            except ValueError:
                d = None
            if d and d > today:
                days_left = (d - today).days
                months_left = max(1, int(days_left / 30))
                per_month = max(0, (target - saved) // months_left)
                line += f"\n   Deadline: {deadline} (jane {days_left} kun)\n   Ayina shama menen: {per_month} som"
            else:
                line += f"\n   Deadline: {deadline}"
        lines.append(line)
    lines.append("\nMaqset ushin jamgarma qosiw: /addgoalsum ID SUMMA yaki `+200000 goal telefon`")
    await update.message.reply_text("\n".join(lines))


async def addgoalsum_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Paydalaniw: /addgoalsum ID SUMMA\nMasalan: /addgoalsum 1 200000"
        )
        return
    try:
        gid = int(context.args[0])
        amount = int(context.args[1])
    except ValueError:
        await update.message.reply_text("ID ham SUMMA putun san boliwi kerak.")
        return
    ok = add_to_goal_db(user.id, wallet_id, gid, amount)
    if not ok:
        await update.message.reply_text("Bunday ID li maqset tawilmadi.")
        return
    await update.message.reply_text("Maqset ushin jamgarma qosildi ✅")


# --- Debts ---

async def lend_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Paydalaniw: /lend SUMMA Ism [YYYY-MM-DD]\n"
            "Misalga: /lend 100000 Ali 2025-01-10"
        )
        return
    try:
        amount = int(context.args[0])
    except ValueError:
        await update.message.reply_text("ID ham SUMMA putun san boliwi kerak.")
        return
    if len(context.args) >= 3:
        name = " ".join(context.args[1:-1])
        due_date = context.args[-1]
    else:
        name = context.args[1]
        due_date = None
    add_debt_db(user.id, wallet_id, "to_me", name, amount, due_date)
    await update.message.reply_text(
        f"Qarz jazildi ✅\nSiz {name} ga {amount} som berdiniz.\nQaytariw kuni: {due_date or 'joq'}"
    )


async def borrow_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Paydalaniw: /borrow SUMMA Ism [YYYY-MM-DD]\n"
            "Masalga: /borrow 100000 Ali 2025-01-10"
        )
        return
    try:
        amount = int(context.args[0])
    except ValueError:
        await update.message.reply_text("SUMMA putun san boliwi kerak.")
        return
    if len(context.args) >= 3:
        name = " ".join(context.args[1:-1])
        due_date = context.args[-1]
    else:
        name = context.args[1]
        due_date = None
    add_debt_db(user.id, wallet_id, "from_me", name, amount, due_date)
    await update.message.reply_text(
        f"Qarz jazildi ✅\nSiz {name} dan {amount} som aldiniz.\nQaytariw kuni: {due_date or 'joq'}"
    )


async def debts_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    rows = list_debts_db(user.id, wallet_id)
    if not rows:
        await update.message.reply_text("Aktiv qarzlar joq ✅")
        return
    lines = ["FAktiv qarzlar:"]
    for did, direction, name, total, remaining, due_date in rows:
        if direction == "to_me":
            dir_text = "Magan qarzdar"
        else:
            dir_text = "Men qarzdarman"
        lines.append(
            f"ID {did}: {name} – {dir_text}\n"
            f"  Jami: {total} som, Qalgan: {remaining} som"
            + (f", Muddeti: {due_date}" if due_date else "")
        )
    lines.append("\nShala qaplaw: /paydebt ID SUMMA")
    await update.message.reply_text("\n".join(lines))


async def paydebt_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if len(context.args) < 2:
        await update.message.reply_text(
            "Paydalaniw: /paydebt ID SUMMA\nMasalga: /paydebt 1 50000"
        )
        return
    try:
        did = int(context.args[0])
        amount = int(context.args[1])
    except ValueError:
        await update.message.reply_text("ID ham SUMMA putun san boliwi kerak.")
        return
    ok, remaining = pay_debt_db(user.id, wallet_id, did, amount)
    if not ok:
        await update.message.reply_text("Bunday ID li qarz tawilmadi.")
        return
    await update.message.reply_text(f"Qarz derlik jawildi ✅\nQalgan summa: {remaining} som")


async def try_handle_debt_payment_text(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, wallet_id: int, text: str):
    lower = text.lower()
    keywords = ["qarz", "kredit", "bank", "toledim", "toledim"]
    if not any(k in lower for k in keywords):
        return False

    m = re.search(r"\d+", text)
    if not m:
        return False
    amount = int(m.group())

    debts = list_debts_db(user_id, wallet_id)
    if not debts:
        return False

    matched = []
    for did, direction, name, total, remaining, due_date in debts:
        if name.lower() in lower:
            matched.append((did, name, direction))

    if len(matched) == 0:
        if len(debts) == 1:
            did, direction, name, total, remaining, due_date = debts[0]
        else:
            await update.message.reply_text(
                "Qarz tolewge uxsas text taptim, biraq qaysi qarzga tiyisli ekenligin aniqlay almadim.\n"
                "/debts orqali ID in korip, /paydebt ID SUMMA komandasinan paydalanin."
            )
            return True
    else:
        did, name, direction = matched[0]

    ok, remaining = pay_debt_db(user_id, wallet_id, did, amount)
    if not ok:
        await update.message.reply_text("Qarz tolewdi jazip aliwda qatelik. /debts di tekserin.")
        return True

    note = f"{name} boynisha qariz tolemi"
    cat_id = add_transaction_db(user_id, wallet_id, amount, "expense", "qarz", note)
    await update.message.reply_text(
        f"Qariz tolemi jazip alindi ✅\n{name} uchun {amount} som tolediniz.\nQalgan summa: {remaining} som"
    )
    await check_budget_alert_and_reply(update.message.reply_text, user_id, wallet_id, cat_id, "qarz")
    return True


# --- Templates / recurring ---

async def addtemplate_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return

    if len(context.args) < 3:
        await update.message.reply_text(
            "Paydalaniw: /addtemplate NOMI SUMMA KATEGORIYA [izoh]\n"
            "Misalga: /addtemplate tuslik 5000 awqat tuste"
        )
        return

    name = context.args[0]
    try:
        amount = int(context.args[1])
    except ValueError:
        await update.message.reply_text("SUMMA putun san boliwi kerek.")
        return
    category = context.args[2].lower()
    note = " ".join(context.args[3:]) if len(context.args) > 3 else ""
    ttype = "expense"
    add_template_db(user.id, wallet_id, name, amount, ttype, category, note)
    await update.message.reply_text(f"Shablon qosildi ✅\nAti: {name}, {amount} som, {category}")


async def templates_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    rows = list_templates_db(user.id, wallet_id)
    if not rows:
        await update.message.reply_text(
            "Shablonlar joq.\nJaratiw: /addtemplate ATI SUMMA KATEGORIYA [sebep]"
        )
        return
    lines = ["Shablonlar:"]
    for tid, name, amount, ttype, category, note in rows:
        sign = "+" if ttype == "income" else "-"
        lines.append(
            f"ID {tid}: {name} – {sign}{amount} som, {category}"
            + (f" ({note})" if note else "")
        )
    lines.append("\nJumsaw: /usetemplate ID")
    await update.message.reply_text("\n".join(lines))


async def usetemplate_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id_default = get_default_wallet_id(user.id)
    if wallet_id_default is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if not context.args:
        await update.message.reply_text("Paydalaniw: /usetemplate ID")
        return
    try:
        tid = int(context.args[0])
    except ValueError:
        await update.message.reply_text("ID ham SUMMA putun san boliwi kerak.")
        return
    row = get_template_db(user.id, wallet_id_default, tid)
    if not row:
        await update.message.reply_text("Bunday ID li shablon tawilmadi.")
        return
    _, name, amount, ttype, category, note = row
    await create_transaction_with_wallet_choice(update, context, user.id, amount, ttype, category, note or name)


async def addrecurring_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyon tawilmadi.")
        return

    if len(context.args) < 3:
        await update.message.reply_text(
            "Paydalaniw: /addrecurring SUMMA KUN KATEGORIYA [tusindirmesi]\n"
            "Masalan: /addrecurring 800000 5 ijara uy"
        )
        return
    try:
        amount = int(context.args[0])
        day_of_month = int(context.args[1])
    except ValueError:
        await update.message.reply_text("SUMMA ham KUN putun san boliwi kerak.")
        return
    category = context.args[2].lower()
    note = " ".join(context.args[3:]) if len(context.args) > 3 else ""
    ttype = "expense"
    if not (1 <= day_of_month <= 28):
        await update.message.reply_text("KUN 1 den 28 shekem boliwi kerek (ay uchun).")
        return
    add_recurring_db(user.id, wallet_id, amount, ttype, category, note, day_of_month)
    await update.message.reply_text(
        f"Qaytalaniwshi tolem qosildi ✅\nHar ay {day_of_month}-kuni {amount} som, {category}"
    )


async def recurring_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyon topilmadi.")
        return
    rows = list_recurring_db(user.id, wallet_id)
    if not rows:
        await update.message.reply_text(
            "Takroriy to'lovlar yo'q.\nYaratish: /addrecurring SUMMA KUN KATEGORIYA [izoh]"
        )
        return
    lines = ["Takroriy to'lovlar:"]
    for rid, amount, ttype, category, note, day_of_month, active in rows:
        status = "faol" if active else "o'chirilgan"
        lines.append(
            f"ID {rid}: Har oy {day_of_month}-kuni "
            f"{'-' if ttype=='expense' else '+'}{amount} so'm, {category}"
            + (f" ({note})" if note else "")
            + f" – {status}"
        )
    lines.append("\nO'chirish: /delrecurring ID")
    await update.message.reply_text("\n".join(lines))


async def delrecurring_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    if not context.args:
        await update.message.reply_text("Paydalaniw: /delrecurring ID")
        return
    try:
        rid = int(context.args[0])
    except ValueError:
        await update.message.reply_text("ID putun san boliwi kerek.")
        return
    ok = deactivate_recurring_db(user.id, wallet_id, rid)
    if not ok:
        await update.message.reply_text("Bunday ID li qaytalaniwshi tolew tawilmadi.")
    else:
        await update.message.reply_text("Qaytalaniwshi tolew oshirildi ✅")


# --- Export / email ---

async def exportcsv_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    csv_mem = export_transactions_csv(user.id, wallet_id)
    today = date.today().isoformat()
    filename = f"transactions_{today}.csv"
    await update.message.reply_document(
        document=csv_mem,
        filename=filename,
        caption="Tranzaksiyalar CSV fayli 📄",
    )


async def exportxlsx_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return
    xlsx_mem = export_transactions_excel(user.id, wallet_id)
    today = date.today().isoformat()
    filename = f"transactions_{today}.xlsx"
    await update.message.reply_document(
        document=xlsx_mem,
        filename=filename,
        caption="Tranzaksiyalar Excel fayli 📊",
    )


async def setemail_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    if not context.args:
        await update.message.reply_text("Paydalaniw: /setemail email@example.com")
        return
    email = context.args[0]
    if "@" not in email or "." not in email:
        await update.message.reply_text("Email manzilidi qate jazganga uqsaysiz")
        return
    set_email_db(user.id, email)
    await update.message.reply_text(f"Email saqlandi ✅\n{email}")


def send_email_with_attachments(to_email: str, subject: str, body: str, attachments):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = FROM_EMAIL
    msg["To"] = to_email
    msg.set_content(body)

    for filename, data, mime_type in attachments:
        maintype, subtype = mime_type.split("/", 1)
        msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=filename)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)


async def emailreport_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi.")
        return

    email = get_email_db(user.id)
    if not email:
        await update.message.reply_text(
            "Aldin email manzildi sazlan Misaldagiday formatta jiberin: /setemail email@example.com"
        )
        return

    fmt = "both"
    if context.args:
        fmt = context.args[0].lower()
        if fmt not in ("csv", "xlsx", "both"):
            await update.message.reply_text("Format qate. csv / xlsx / both dan birewin tanlan.")
            return

    attachments = []
    today = date.today().isoformat()
    if fmt in ("csv", "both"):
        csv_mem = export_transactions_csv(user.id, wallet_id)
        attachments.append((f"transactions_{today}.csv", csv_mem.getvalue(), "text/csv"))
    if fmt in ("xlsx", "both"):
        xlsx_mem = export_transactions_excel(user.id, wallet_id)
        attachments.append(
            (
                f"transactions_{today}.xlsx",
                xlsx_mem.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )

    try:
        send_email_with_attachments(
            to_email=email,
            subject="Finans esabat fayllari",
            body="Finans tranzaksiyalariniz esabat fayllari dasturge jiberildi.",
            attachments=attachments,
        )
        await update.message.reply_text(f"Esabat email arqali jiberildi ✅\n{email}")
    except Exception as e:
        logger.error(f"Email send error: {e}")
        await update.message.reply_text("Email jiberiwde qatelik juz berdi. SMTP sazlamalarin teksirin.")


# --- Settings ---

async def settings_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT daily_reminder_enabled, weekly_report_enabled, email, daily_time, weekly_time, language FROM users WHERE user_id = ?",
        (user.id,),
    )
    row = cur.fetchone()
    conn.close()
    daily = bool(row[0]) if row else True
    weekly = bool(row[1]) if row else True
    email = row[2] if row else None
    d_time = row[3] or "22:00"
    w_time = row[4] or "20:00"
    lang = row[5] or "UZB"

    lines = [
        "⚙️ Sazlamalar:",
        f"- Kunlik eslatpe: {'jagilgan' if daily else 'oshirilgen'} (/toggledaily)",
        f"  Waqit: {d_time} (/setdailytime HH:MM)",
        f"- Haptelik esabat: {'jogilgan' if weekly else 'oshirilgen'} (/toggleweekly)",
        f"  Waqit (Yekshembi): {w_time} (/setweeklytime HH:MM)",
        f"- Email: {email or 'ornatilmagan'} (/setemail email@example.com)",
        f"- Til: {lang} (/lang arqali ozgerttiriw)",
    ]
    await update.message.reply_text("\n".join(lines))


async def toggledaily_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT daily_reminder_enabled FROM users WHERE user_id = ?",
        (user.id,),
    )
    row = cur.fetchone()
    current = bool(row[0]) if row else True
    new_val = 0 if current else 1
    cur.execute(
        "UPDATE users SET daily_reminder_enabled = ? WHERE user_id = ?",
        (new_val, user.id),
    )
    conn.commit()
    conn.close()
    await update.message.reply_text(
        f"Kunlik esletpe {'oshirilgen' if new_val == 0 else 'jagilgan'} ✅"
    )


async def toggleweekly_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT weekly_report_enabled FROM users WHERE user_id = ?",
        (user.id,),
    )
    row = cur.fetchone()
    current = bool(row[0]) if row else True
    new_val = 0 if current else 1
    cur.execute(
        "UPDATE users SET weekly_report_enabled = ? WHERE user_id = ?",
        (new_val, user.id),
    )
    conn.commit()
    conn.close()
    await update.message.reply_text(
        f"Haptelik esabat {'oshirilgen' if new_val == 0 else 'jagilgan'} ✅"
    )


async def setdailytime_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    if not context.args:
        await update.message.reply_text("Paydalaniw: /setdailytime HH:MM\nMisalga: /setdailytime 22:30")
        return
    tstr = context.args[0]
    t = parse_time_str(tstr)
    if t is None:
        await update.message.reply_text("Waqit formati qate. HH:MM (24 sagatliq format).")
        return
    set_daily_time_db(user.id, tstr)
    schedule_daily_job(context.application.job_queue, user.id, tstr)
    await update.message.reply_text(f"Kunlik esletpe waqti {tstr} ga ozgerttirildi ✅")


async def setweeklytime_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    if not context.args:
        await update.message.reply_text("Paydalaniw: /setweeklytime HH:MM\nMisalga: /setweeklytime 20:00")
        return
    tstr = context.args[0]
    t = parse_time_str(tstr)
    if t is None:
        await update.message.reply_text("Waqit formati qate. HH:MM (24 sagatliq format).")
        return
    set_weekly_time_db(user.id, tstr)
    schedule_weekly_job(context.application.job_queue, user.id, tstr)
    await update.message.reply_text(f"Haftalik esabat waqti {tstr} ga ozgerttirildi ✅")


# --- Quick kategoriya callback ---

async def category_quick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    data = query.data
    _, key = data.split(":", 1)
    title = quick_category_title(key)
    set_pending_category_db(user_id, key)
    await query.message.reply_text(
        f"{title} kategoriya tanlandi.\nEndi summani jazin, misalga: 50000"
    )


async def create_transaction_with_wallet_choice(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                                user_id: int, amount: int, ttype: str, category: str, note: str):
    wallets = list_wallets_db(user_id)
    if len(wallets) <= 1:
        wallet_id = get_default_wallet_id(user_id)
        cat_id = add_transaction_db(user_id, wallet_id, amount, ttype, category, note)
        sign = "+" if ttype == "income" else "-"
        await update.message.reply_text(
            f"Jazip aldim ✅\n{sign}{amount} som | {category}"
            + (f" | {note}" if note else "")
        )
        await check_budget_alert_and_reply(update.message.reply_text, user_id, wallet_id, cat_id, category)
        return

    pid = create_pending_transaction_db(user_id, amount, ttype, category, note)
    keyboard = []
    for wid, name, currency, is_default in wallets:
        label = f"{name} [{currency}]"
        if is_default:
            label += " (hazirgi)"
        keyboard.append(
            [InlineKeyboardButton(label, callback_data=f"pt:{pid}:{wid}")]
        )
    await update.message.reply_text(
        "Qaysi hamyanga jazamiz?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def pending_transaction_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data  # pt:pending_id:wallet_id
    _, pid_str, wid_str = data.split(":")
    try:
        pid = int(pid_str)
        wid = int(wid_str)
    except ValueError:
        await query.message.reply_text("Qatelik: qate magliwmat.")
        return
    user_id = query.from_user.id
    row = get_pending_transaction_db(user_id, pid)
    if not row:
        await query.message.reply_text("Bul tranzaksiya muddeti otgan yaki tawilmadi.")
        return
    _, _, amount, ttype, category, note = row
    cat_id = add_transaction_db(user_id, wid, amount, ttype, category, note)
    delete_pending_transaction_db(pid)
    sign = "+" if ttype == "income" else "-"
    await query.message.reply_text(
        f"Jazip aldim ✅\n{sign}{amount} som | {category}"
        + (f" | {note}" if note else "")
    )
    await check_budget_alert_and_reply(query.message.reply_text, user_id, wid, cat_id, category)


# --- Tiykargi text handler ---

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user.id)
    wallet_id = get_default_wallet_id(user.id)
    if wallet_id is None:
        await update.message.reply_text("Hamyan tawilmadi. /start ti basip korin.")
        return

    text = update.message.text.strip()

    # Tomendegi tuymelerdi emodji boyinsha tanlaw:
    if text.startswith("📊"):
        await today_cmd(update, context)
        return
    elif text.startswith("📆"):
        await month_cmd(update, context)
        return
    elif text.startswith("🎯"):
        await goals_cmd(update, context)
        return
    elif text.startswith("⚙️") or text.startswith("⚙"):
        await settings_cmd(update, context)
        return
    elif text.startswith("➕"):
        await update.message.reply_text(
            "Qarejet yaki daramatti tomendegishe jazin yaki kategoriya tuymesin tanlan:\n"
            "`-50000 awqat`\n"
            "`+200000 ayliq`\n"
            "`bugun 120000 bazar`",
            parse_mode="Markdown",
            reply_markup=quick_categories_keyboard(),
        )
        return

    handled_debt = await try_handle_debt_payment_text(update, context, user.id, wallet_id, text)
    if handled_debt:
        return

    goal_parsed = parse_goal_contribution_text(text)
    if goal_parsed:
        amount, goal_name_part = goal_parsed
        row = find_goal_by_name(user.id, wallet_id, goal_name_part)
        if not row:
            await update.message.reply_text("Bunday atqa uxsas maqset tawilmadi.")
            return
        gid, name, target, saved = row
        ok = add_to_goal_db(user.id, wallet_id, gid, amount)
        if ok:
            await update.message.reply_text(
                f"Maqset qosildi ✅\n{name}: +{amount} som"
            )
            return

    pending_cat = get_pending_category_db(user.id)

    parsed = parse_transaction_text(text)
    if not parsed:
        if pending_cat and text.lstrip("+-").isdigit():
            amount = int(text)
            if amount < 0:
                amount = -amount
            ttype = "expense"
            category = pending_cat
            note = ""
            set_pending_category_db(user.id, None)
            await create_transaction_with_wallet_choice(update, context, user.id, amount, ttype, category, note)
            return

        await update.message.reply_text(
            "Tusinbedim 🤔\nMisal:\n"
            "`-50000 awqat`\n"
            "`+200000 ayliq`\n"
            "`bugun 120000 bazar`",
            parse_mode="Markdown",
        )
        return

    amount, ttype, category, note = parsed

    if pending_cat and category == "bashqa" and not note and ttype == "expense":
        category = pending_cat
    set_pending_category_db(user.id, None)

    await create_transaction_with_wallet_choice(update, context, user.id, amount, ttype, category, note)


# ===================== JOBLAR =====================

async def daily_reminder_job(context: ContextTypes.DEFAULT_TYPE):
    job = context.job
    user_id = job.data["user_id"]
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT daily_reminder_enabled FROM users WHERE user_id = ?",
        (user_id,),
    )
    row = cur.fetchone()
    conn.close()
    if not row or not bool(row[0]):
        return
    try:
        await context.bot.send_message(
            chat_id=user_id,
            text="Kun juwmaqlandi 😊\nBugungi daramat ham shiginlarinizdi botga jaziwdi umitpan.",
        )
    except Exception as e:
        logger.warning(f"Daily reminder send error to {user_id}: {e}")


async def weekly_report_job(context: ContextTypes.DEFAULT_TYPE):
    job = context.job
    user_id = job.data["user_id"]
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT weekly_report_enabled FROM users WHERE user_id = ?",
        (user_id,),
    )
    row = cur.fetchone()
    conn.close()
    if not row or not bool(row[0]):
        return

    today = date.today()
    if today.weekday() != 6:
        return

    wallet_id = get_default_wallet_id(user_id)
    if wallet_id is None:
        return
    start_dt, end_dt = get_week_bounds(today)
    income, expense, cat_rows = get_period_stats(user_id, wallet_id, start_dt, end_dt)
    balance = (income or 0) - (expense or 0)
    text = (
        "📊 Haptelik qisqa esabat:\n\n"
        f"Daramat: {income or 0} som\n"
        f"Shigim: {expense or 0} som\n"
        f"Balans: {balance} som\n"
    )
    if expense > 0 and cat_rows:
        text += "\nEn kop shigim bolgan kategoriyalar:\n"
        for name, total in cat_rows[:3]:
            percent = int((total * 100) / expense)
            text += f"  • {name}: {total} som ({percent}%)\n"
    try:
        await context.bot.send_message(chat_id=user_id, text=text)
    except Exception as e:
        logger.warning(f"Weekly report send error to {user_id}: {e}")


async def recurring_payments_job(context: ContextTypes.DEFAULT_TYPE):
    logger.info("Recurring payments job isledi")
    today = date.today()
    day = today.day
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, user_id, wallet_id, amount, type, category, note
        FROM recurring_payments
        WHERE active = 1 AND day_of_month = ?
        """,
        (day,),
    )
    rows = cur.fetchall()
    conn.close()
    for rid, user_id, wallet_id, amount, ttype, category, note in rows:
        cat_id = add_transaction_db(user_id, wallet_id, amount, ttype, category, note)
        msg = (
            f"Qaytalaniwshi tolem amelge asirildi ✅\n"
            f"{'-' if ttype=='expense' else '+'}{amount} som | {category}"
        )
        try:
            await context.bot.send_message(chat_id=user_id, text=msg)
        except Exception as e:
            logger.warning(f"Recurring send error to {user_id}: {e}")
        async def dummy_reply(text):
            return None
        await check_budget_alert_and_reply(dummy_reply, user_id, wallet_id, cat_id, category)


async def debt_reminder_job(context: ContextTypes.DEFAULT_TYPE):
    logger.info("Debt reminder job isledi")
    today = date.today()
    target_dates = {
        today + timedelta(days=3): "3 kunnen keyin qaytariw muddeti keladi.",
        today: "Bugun qaytariw muddeti.",
    }
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, user_id, wallet_id, direction, counterparty_name,
               total_amount, remaining_amount, due_date
        FROM debts
        WHERE remaining_amount > 0 AND due_date IS NOT NULL
        """
    )
    rows = cur.fetchall()
    conn.close()
    for did, user_id, wallet_id, direction, name, total, remaining, due_date in rows:
        try:
            d = datetime.fromisoformat(due_date).date()
        except Exception:
            continue
        if d in target_dates:
            if direction == "to_me":
                text = (
                    f"Qarz eslatpesi 📌\n"
                    f"{name} sizge {remaining} som qaytariwi kerak.\n"
                    f"Muddet: {due_date}\n"
                    f"{target_dates[d]}"
                )
            else:
                text = (
                    f"Qarz esletpe 📌\n"
                    f"Siz {name} ga {remaining} som qaytariwiniz kerek.\n"
                    f"Muddet: {due_date}\n"
                    f"{target_dates[d]}"
                )
            try:
                await context.bot.send_message(chat_id=user_id, text=text)
            except Exception as e:
                logger.warning(f"Debt reminder error to {user_id}: {e}")


# --- Unknown command ---

async def unknown_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Bul komandani bilmiymen. /help di korip shigin 🙂")


# ===================== WEB ADMIN PANEL =====================

def is_logged_in():
    return session.get("logged_in") is True


@admin_app.route("/")
def admin_index():
    if not is_logged_in():
        return redirect(url_for("admin_login"))
    return redirect(url_for("admin_dashboard"))


@admin_app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_WEB_USERNAME and password == ADMIN_WEB_PASSWORD:
            session["logged_in"] = True
            return redirect(url_for("admin_dashboard"))
        else:
            error = "Login yaki parol qate."
    # Endi HTML ni fayldan olamiz
    return render_template("admin_login.html", error=error)

    html = """
    <html>
    <head><title>Finance Bot Admin - Login</title></head>
    <body>
      <h1>Finance Bot Admin - Login</h1>
      {% if error %}
        <p style="color:red;">{{ error }}</p>
      {% endif %}
      <form method="post">
        <label>Login:
          <input type="text" name="username">
        </label>
        <br>
        <label>Parol:
          <input type="password" name="password">
        </label>
        <br><br>
        <button type="submit">Kirish</button>
      </form>
    </body>
    </html>
    """
    return render_template_string(html, error=error)


@admin_app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@admin_app.route("/admin")
def admin_dashboard():
    if not is_logged_in():
        return redirect(url_for("admin_login"))

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    filter_user_id = request.args.get("user_id", "")

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM users")
    user_count = cur.fetchone()[0] or 0
    cur.execute("SELECT COUNT(*) FROM transactions")
    tx_count = cur.fetchone()[0] or 0

    where_clauses = []
    params = []

    if from_date:
        where_clauses.append("datetime(t.created_at) >= ?")
        params.append(from_date + " 00:00:00")
    if to_date:
        where_clauses.append("datetime(t.created_at) <= ?")
        params.append(to_date + " 23:59:59")
    if filter_user_id:
        try:
            uid = int(filter_user_id)
            where_clauses.append("t.user_id = ?")
            params.append(uid)
        except ValueError:
            filter_user_id = ""

    base_sql = """
        SELECT t.created_at, t.type, t.amount, c.name, t.note, w.name, t.user_id
        FROM transactions t
        JOIN categories c ON t.category_id = c.id
        JOIN wallets w ON t.wallet_id = w.id
    """
    if where_clauses:
        base_sql += " WHERE " + " AND ".join(where_clauses)
    base_sql += " ORDER BY datetime(t.created_at) DESC LIMIT 50"

    cur.execute(base_sql, params)
    last_tx = cur.fetchall()
    conn.close()

    return render_template(
        "admin_dashboard.html",
        user_count=user_count,
        tx_count=tx_count,
        last_tx=last_tx,
        from_date=from_date,
        to_date=to_date,
        filter_user_id=filter_user_id,
    )

@admin_app.route("/admin/tx_chart.png")
def tx_chart():
    if not is_logged_in():
        return redirect(url_for("admin_login"))

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    filter_user_id = request.args.get("user_id", "")

    where_clauses = []
    params = []

    if from_date:
        where_clauses.append("datetime(created_at) >= ?")
        params.append(from_date + " 00:00:00")
    if to_date:
        where_clauses.append("datetime(created_at) <= ?")
        params.append(to_date + " 23:59:59")
    if filter_user_id:
        try:
            uid = int(filter_user_id)
            where_clauses.append("user_id = ?")
            params.append(uid)
        except ValueError:
            pass

    conn = get_conn()
    cur = conn.cursor()
    sql = "SELECT date(created_at), COUNT(*) FROM transactions"
    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    sql += " GROUP BY date(created_at) ORDER BY date(created_at)"
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    dates = [r[0] for r in rows]
    counts = [r[1] for r in rows]

    fig, ax = plt.subplots()
    if dates:
        ax.plot(dates, counts, marker="o")
    ax.set_title("Kunlik tranzaksiyalar sani")
    ax.set_xlabel("Sane")
    ax.set_ylabel("San")
    fig.autofmt_xdate()

    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)

    return send_file(buf, mimetype="image/png")


@admin_app.route("/admin/category_chart.png")
def category_chart():
    if not is_logged_in():
        return redirect(url_for("admin_login"))

    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    filter_user_id = request.args.get("user_id", "")

    where_clauses = ["t.type = 'expense'"]
    params = []

    if from_date:
        where_clauses.append("datetime(t.created_at) >= ?")
        params.append(from_date + " 00:00:00")
    if to_date:
        where_clauses.append("datetime(t.created_at) <= ?")
        params.append(to_date + " 23:59:59")
    if filter_user_id:
        try:
            uid = int(filter_user_id)
            where_clauses.append("t.user_id = ?")
            params.append(uid)
        except ValueError:
            pass

    conn = get_conn()
    cur = conn.cursor()
    sql = """
        SELECT c.name, SUM(t.amount)
        FROM transactions t
        JOIN categories c ON t.category_id = c.id
    """
    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    sql += " GROUP BY c.name ORDER BY SUM(t.amount) DESC"
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    labels = [r[0] for r in rows]
    sizes = [r[1] for r in rows]

    if not labels:
        fig, ax = plt.subplots()
        ax.text(0.5, 0.5, "Magliwmat joq", ha="center", va="center")
        ax.axis("off")
    else:
        fig, ax = plt.subplots()
        ax.pie(sizes, labels=labels, autopct="%1.1f%%")
        ax.set_title("Kategoriyalar boyinsha shigimlar")

    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)

    return send_file(buf, mimetype="image/png")


def run_admin_panel():
    admin_app.run(host="0.0.0.0", port=8000)


# ===================== MAIN =====================

def main():
    init_db()

    app = ApplicationBuilder().token(TELEGRAM_BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start_cmd))
    app.add_handler(CommandHandler("help", help_cmd))

    app.add_handler(CommandHandler("lang", lang_cmd))
    app.add_handler(CallbackQueryHandler(lang_callback, pattern=r"^lang:"))

    app.add_handler(CommandHandler("wallets", wallets_cmd))
    app.add_handler(CommandHandler("addwallet", addwallet_cmd))
    app.add_handler(CommandHandler("setwallet", setwallet_cmd))
    app.add_handler(CommandHandler("totalwealth", totalwealth_cmd))

    app.add_handler(CommandHandler("today", today_cmd))
    app.add_handler(CommandHandler("week", week_cmd))
    app.add_handler(CommandHandler("month", month_cmd))
    app.add_handler(CommandHandler("year", year_cmd))
    app.add_handler(CommandHandler("report", report_cmd))
    app.add_handler(CallbackQueryHandler(report_callback, pattern=r"^report:"))

    app.add_handler(CommandHandler("categories", categories_cmd))
    app.add_handler(CommandHandler("setbudget", setbudget_cmd))
    app.add_handler(CommandHandler("budget", budget_cmd))

    app.add_handler(CommandHandler("addgoal", addgoal_cmd))
    app.add_handler(CommandHandler("goals", goals_cmd))
    app.add_handler(CommandHandler("addgoalsum", addgoalsum_cmd))

    app.add_handler(CommandHandler("lend", lend_cmd))
    app.add_handler(CommandHandler("borrow", borrow_cmd))
    app.add_handler(CommandHandler("debts", debts_cmd))
    app.add_handler(CommandHandler("paydebt", paydebt_cmd))

    app.add_handler(CommandHandler("addtemplate", addtemplate_cmd))
    app.add_handler(CommandHandler("templates", templates_cmd))
    app.add_handler(CommandHandler("usetemplate", usetemplate_cmd))

    app.add_handler(CommandHandler("addrecurring", addrecurring_cmd))
    app.add_handler(CommandHandler("recurring", recurring_cmd))
    app.add_handler(CommandHandler("delrecurring", delrecurring_cmd))

    app.add_handler(CommandHandler("exportcsv", exportcsv_cmd))
    app.add_handler(CommandHandler("exportxlsx", exportxlsx_cmd))
    app.add_handler(CommandHandler("setemail", setemail_cmd))
    app.add_handler(CommandHandler("emailreport", emailreport_cmd))

    app.add_handler(CommandHandler("settings", settings_cmd))
    app.add_handler(CommandHandler("toggledaily", toggledaily_cmd))
    app.add_handler(CommandHandler("toggleweekly", toggleweekly_cmd))
    app.add_handler(CommandHandler("setdailytime", setdailytime_cmd))
    app.add_handler(CommandHandler("setweeklytime", setweeklytime_cmd))

    app.add_handler(CallbackQueryHandler(category_quick_callback, pattern=r"^cat:"))
    app.add_handler(CallbackQueryHandler(pending_transaction_callback, pattern=r"^pt:"))

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    # Axirinda belgisiz komandalar:
    app.add_handler(MessageHandler(filters.COMMAND, unknown_cmd))

    jq = app.job_queue
    schedule_all_users_jobs(jq)
    jq.run_daily(
        recurring_payments_job,
        time=time(hour=6, minute=0),
        name="recurring_payments",
    )
    jq.run_daily(
        debt_reminder_job,
        time=time(hour=9, minute=0),
        name="debt_reminder",
    )

    threading.Thread(target=run_admin_panel, daemon=True).start()

    app.run_polling()


if __name__ == "__main__":
    main()
